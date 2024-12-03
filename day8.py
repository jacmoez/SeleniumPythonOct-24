from selenium import webdriver
from openpyxl.styles import Font, PatternFill, Alignment
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import openpyxl
import time

def report_excel(file_name="report.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Report"
    ws.append(["Test", "Message"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="ffffff")
        cell.fill = PatternFill(start_color="42c328", end_color="42c328", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        wb.save(file_name)

def write_report_excel(test="",msg="", file_name="report.xlsx"):
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    ws.append([test,msg])
    wb.save(file_name)


class SwagTest:

    driver = webdriver.Firefox()

    def open_browser(self):
        self.driver.get("https://www.saucedemo.com/")
        write_report_excel("Test 1:"," Open browser Success")


    def login_test(self):
        self.driver.find_element(By.ID, "user-name").send_keys("standard_user")
        time.sleep(2)
        self.driver.find_element(By.ID, "password").send_keys("secret_sauce", Keys.ENTER)
        write_report_excel("Test 2:"," Login Success")
        time.sleep(2)


    def check_page_go(self):
        ps = self.driver.page_source
        write_report_excel("Test 3 :"," inventory Page Arrived " if "Products" in ps else "Wrong Page")

    def add_to_cart(self):
        self.driver.find_element(By.ID, "add-to-cart-sauce-labs-backpack").click()
        time.sleep(2)
        self.driver.find_element(By.ID, "add-to-cart-test.allthethings()-t-shirt-(red)").click()
        time.sleep(2)
        self.driver.find_element(By.ID, "add-to-cart-sauce-labs-onesie").click()
        time.sleep(2)
        write_report_excel("Test 4 :"," Add Item to Cart")


    def remove_item(self):
        self.driver.find_element(By.ID, "remove-sauce-labs-backpack").click()
        time.sleep(2)
        write_report_excel("Test 5 :"," Remove 1 item")

    def view_order_item(self):
        self.driver.find_element(By.CLASS_NAME, "shopping_cart_link").click()
        time.sleep(2)
        write_report_excel("Test 6 :"," View Order Item")

    def check_cart_page_go(self):
        ps = self.driver.page_source
        time.sleep(2)
        write_report_excel("Test 7 :"," Cart Page Arrived " if "Your Cart" in ps else "Wrong Page!")

    def remove_item_from_cart(self):
        self.driver.find_element(By.ID, "remove-sauce-labs-onesie").click()
        time.sleep(2)
        write_report_excel("Test 8:"," Remove item from cart - Cart Page")

    def check_out(self):
        self.driver.find_element(By.ID, "checkout").click()
        write_report_excel("Test 9 :"," Click Check Out Button")

    def check_checkout_page_go(self):
        ps = self.driver.page_source
        time.sleep(2)
        write_report_excel("Test 10 :"," Checkout Page Arrived " if "Checkout: Your Information" in ps else "Wrong Page!")

    def check_out_information(self):
        self.driver.find_element(By.ID, "first-name").send_keys("QA")
        time.sleep(2)
        self.driver.find_element(By.ID, "last-name").send_keys("Testing")
        time.sleep(2)
        self.driver.find_element(By.ID, "postal-code").send_keys("11111")
        time.sleep(2)
        self.driver.find_element(By.ID, "continue").click()
        write_report_excel("Test 11 :"," Checkout Continue Success!")

    def check_payment(self):
        item_total = self.driver.find_element(By.CLASS_NAME, "summary_subtotal_label").text
        tax_total = self.driver.find_element(By.CLASS_NAME, "summary_tax_label").text
        total_price = self.driver.find_element(By.CLASS_NAME, "summary_total_label").text
        write_report_excel("Item Total : ", item_total)
        write_report_excel("Tax Price : ", tax_total)
        write_report_excel("Total Price : ", total_price)
        write_report_excel("Test 12 :"," Payment Success!")

    def finish(self):
        self.driver.find_element(By.ID, "finish").click()
        time.sleep(2)
        ps = self.driver.page_source
        write_report_excel("Test 13 :"," Finish Page Arrived " if "Thank you for your order!" in ps else "Wrong Page!")


    def main(self):
        report_excel()
        self.open_browser()
        self.login_test()
        self.check_page_go()
        self.add_to_cart()
        self.remove_item()
        self.view_order_item()
        self.check_cart_page_go()
        self.remove_item_from_cart()
        self.check_out()
        self.check_checkout_page_go()
        self.check_out_information()
        self.check_payment()
        self.finish()

SwagTest().main()